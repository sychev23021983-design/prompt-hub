"""
Импортирует данные из трёх xlsx-файлов (semantic-core каждого проекта) в единую SQLite базу.

Запуск: python import.py
Ожидает файлы в /app/import_data/ (или ./import_data/ при локальном запуске):
  - shustrik-maps_SEO_audit_and_core.xlsx
  - itc-by_SEO_audit_and_core.xlsx
  - fit-shustrik-maps_SEO_core.xlsx

Идемпотентно: при повторном запуске очищает и перезаполняет таблицы (данные о
прогрессе — Applied?/Date Applied/Notes — переносятся по ключу URL/Name, чтобы не
терять отметки о сделанной работе при повторном импорте.
"""
import json
import os
import openpyxl
from db import get_conn, init_db

IMPORT_DIR = os.environ.get("PROMPT_HUB_IMPORT_DIR", "/app/import_data")

SITES = [
    {
        "slug": "shustrik-maps",
        "name": "shustrik-maps.com",
        "language": "en",
        "prompt_style": "commerce_en",
        "source_repo": "shustrik-maps-seo",
        "file": "shustrik-maps_SEO_audit_and_core.xlsx",
    },
    {
        "slug": "itc-by",
        "name": "itc.by (ИТЦ-М)",
        "language": "ru",
        "prompt_style": "b2b_ru",
        "source_repo": "itc-by-seo",
        "file": "itc-by_SEO_audit_and_core.xlsx",
    },
    {
        "slug": "fit-shustrik-maps",
        "name": "fit.shustrik-maps.com",
        "language": "ru",
        "prompt_style": "personal_ru",
        "source_repo": "fit-shustrik-maps-seo",
        "file": "fit-shustrik-maps_SEO_core.xlsx",
    },
]


def cell_map(ws):
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, row))


def load_existing_progress(conn):
    """Key = (site_slug, sheet, name) -> (applied, date_applied, notes, url,
    structure_node_id, page_type, page_plan). Все они переносятся при
    переимпорте, чтобы вставленный вручную URL, привязка к разделу структуры,
    выбранный тип страницы и вставленный план страницы не терялись, когда
    xlsx обновляется свежей выгрузкой из KeyCollector.

    Важно: ключ строится по name, а НЕ по url — url теперь редактируемое поле
    (см. EDITABLE_FIELDS в server.py), и если бы ключ зависел от url, повторный
    импорт после ручной правки url не находил бы совпадения (в xlsx URL остался
    прежним) и сбрасывал бы и url, и structure_node_id, и applied/notes."""
    progress = {}
    rows = conn.execute("""
        SELECT s.slug, r.sheet, r.url, r.name, r.applied, r.date_applied, r.notes,
               r.structure_node_id, r.page_type, r.page_plan
        FROM rows_ r JOIN sites s ON s.id = r.site_id
    """).fetchall()
    for r in rows:
        key = (r["slug"], r["sheet"], r["name"])
        progress[key] = (r["applied"], r["date_applied"], r["notes"], r["url"],
                          r["structure_node_id"], r["page_type"], r["page_plan"])
    return progress


def import_shustrik_maps(conn, site_id, path, progress):
    wb = openpyxl.load_workbook(path, data_only=True)

    for row in cell_map(wb["Products_SEO"]):
        key = ("shustrik-maps", "Products_SEO", row.get("Current Name"))
        applied, date_applied, notes, url, node_id, page_type, page_plan = progress.get(
            key, (row.get("Applied?", "No"), row.get("Date Applied"), row.get("Notes"), row.get("URL"), None, None, None))
        conn.execute("""INSERT INTO rows_
            (site_id, sheet, name, url, row_type, seo_title, h1, meta_description,
             primary_keyword, secondary_keywords, lsi_keywords, faq_questions, extra_json,
             applied, date_applied, notes, structure_node_id, page_type, page_plan)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (site_id, "Products_SEO", row.get("Current Name"), url, row.get("Type"),
             row.get("SEO Title"), row.get("H1"), row.get("Meta Description"),
             row.get("Primary Keyword"), row.get("Secondary Keywords"), row.get("LSI Keywords"),
             None, json.dumps({"geo": row.get("Geo"), "geo_level": row.get("Geo Level"),
                                "status": row.get("Status"), "action": row.get("Action")}, ensure_ascii=False),
             applied, date_applied, notes, node_id, page_type, page_plan))

    for row in cell_map(wb["Categories"]):
        key = ("shustrik-maps", "Categories", row.get("Category"))
        applied, date_applied, notes, url, node_id, page_type, page_plan = progress.get(
            key, (row.get("Applied?", "No"), None, row.get("Notes"), row.get("URL"), None, None, None))
        conn.execute("""INSERT INTO rows_
            (site_id, sheet, name, url, row_type, seo_title, h1, meta_description,
             primary_keyword, secondary_keywords, lsi_keywords, extra_json, applied, date_applied, notes,
             structure_node_id, page_type, page_plan)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (site_id, "Categories", row.get("Category"), url, row.get("Group"),
             row.get("SEO Title"), row.get("H1"), row.get("Meta Description"),
             row.get("Primary Keyword"), row.get("Secondary Keywords"), row.get("LSI Keywords"),
             json.dumps({"format": row.get("Format")}, ensure_ascii=False),
             applied, date_applied, notes, node_id, page_type, page_plan))


def import_itc_by(conn, site_id, path, progress):
    wb = openpyxl.load_workbook(path, data_only=True)
    for row in cell_map(wb["Pages_SEO"]):
        key = ("itc-by", "Pages_SEO", row.get("Name"))
        applied, date_applied, notes, url, node_id, page_type, page_plan = progress.get(
            key, (row.get("Applied?", "No"), row.get("Date Applied"), row.get("Notes"), row.get("URL"), None, None, None))
        conn.execute("""INSERT INTO rows_
            (site_id, sheet, name, url, row_type, seo_title, h1, meta_description,
             primary_keyword, secondary_keywords, lsi_keywords, extra_json, applied, date_applied, notes,
             structure_node_id, page_type, page_plan)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (site_id, "Pages_SEO", row.get("Name"), url, row.get("Type"),
             row.get("SEO Title"), row.get("H1"), row.get("Meta Description"),
             row.get("Primary Keyword"), row.get("Secondary Keywords"), row.get("LSI Keywords"),
             json.dumps({"section": row.get("Section"), "vendor": row.get("Vendor")}, ensure_ascii=False),
             applied, date_applied, notes, node_id, page_type, page_plan))


def import_fit(conn, site_id, path, progress):
    wb = openpyxl.load_workbook(path, data_only=True)
    for row in cell_map(wb["Pages_SEO"]):
        key = ("fit-shustrik-maps", "Pages_SEO", row.get("Page / Group"))
        applied, date_applied, notes, url, node_id, page_type, page_plan = progress.get(
            key, (row.get("Applied?", "No"), row.get("Date Applied"), row.get("Notes"), None, None, None, None))
        conn.execute("""INSERT INTO rows_
            (site_id, sheet, name, url, row_type, seo_title, h1, meta_description,
             primary_keyword, secondary_keywords, lsi_keywords, faq_questions, extra_json,
             applied, date_applied, notes, structure_node_id, page_type, page_plan)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (site_id, "Pages_SEO", row.get("Page / Group"), url, row.get("Type"),
             row.get("SEO Title"), row.get("H1"), row.get("Meta Description"),
             row.get("Primary Keyword"), row.get("Secondary Keywords"), None,
             row.get("FAQ Questions"),
             json.dumps({"all_keywords": row.get("All Keywords"), "keyword_count": row.get("Keyword Count")}, ensure_ascii=False),
             applied, date_applied, notes, node_id, page_type, page_plan))

    for row in cell_map(wb["Blog_Clusters"]):
        key = ("fit-shustrik-maps", "Blog_Clusters", row.get("Topic (Article Idea)"))
        applied, date_applied, notes, url, node_id, page_type, page_plan = progress.get(
            key, (row.get("Applied?", "No"), row.get("Date Applied"), row.get("Notes"), None, None, None, None))
        conn.execute("""INSERT INTO rows_
            (site_id, sheet, name, url, row_type, seo_title, h1, meta_description,
             primary_keyword, secondary_keywords, extra_json, applied, date_applied, notes, structure_node_id,
             page_type, page_plan)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (site_id, "Blog_Clusters", row.get("Topic (Article Idea)"), url, "blog_cluster",
             row.get("SEO Title"), row.get("H1"), row.get("Meta Description"),
             row.get("Primary Keyword"), row.get("Secondary Keywords"),
             json.dumps({"sample_keywords": row.get("Sample Keywords"),
                         "keyword_count": row.get("Keyword Count")}, ensure_ascii=False),
             applied, date_applied, notes, node_id, page_type, page_plan))


IMPORTERS = {
    "shustrik-maps": import_shustrik_maps,
    "itc-by": import_itc_by,
    "fit-shustrik-maps": import_fit,
}


def main():
    init_db()
    conn = get_conn()
    progress = load_existing_progress(conn)

    # rows_ можно безопасно перезаполнять (прогресс/URL/привязка к структуре
    # переносятся через progress по ключу URL/название — см. load_existing_progress).
    # sites НЕЛЬЗЯ удалять и пересоздавать: structure_nodes.site_id ссылается на
    # конкретный id сайта, и его смена при реимпорте оторвала бы всю структуру
    # от сайта. Поэтому sites обновляются на месте (upsert), id стабилен.
    conn.execute("DELETE FROM rows_")
    conn.commit()

    for site in SITES:
        existing = conn.execute("SELECT id FROM sites WHERE slug=?", (site["slug"],)).fetchone()
        if existing:
            site_id = existing["id"]
            conn.execute(
                "UPDATE sites SET name=?, language=?, prompt_style=?, source_repo=? WHERE id=?",
                (site["name"], site["language"], site["prompt_style"], site["source_repo"], site_id))
        else:
            cur = conn.execute(
                "INSERT INTO sites (slug, name, language, prompt_style, source_repo) VALUES (?,?,?,?,?)",
                (site["slug"], site["name"], site["language"], site["prompt_style"], site["source_repo"]))
            site_id = cur.lastrowid
        conn.commit()

        path = os.path.join(IMPORT_DIR, site["file"])
        if not os.path.exists(path):
            print(f"! Файл не найден, пропускаю: {path}")
            continue
        IMPORTERS[site["slug"]](conn, site_id, path, progress)
        conn.commit()
        print(f"Импортирован сайт: {site['name']}")

    total = conn.execute("SELECT COUNT(*) c FROM rows_").fetchone()["c"]
    print(f"Готово. Всего строк в базе: {total}")
    conn.close()


if __name__ == "__main__":
    main()
